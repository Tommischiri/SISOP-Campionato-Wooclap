import openpyxl
import argparse
import os
import prettytable

excel_extension = {"xlsx","xlsm","xlsb","xltx","xltm"}

#def absoluteFilePaths(directory):
#    for dirpath,_,filenames in os.walk(directory):
#        for f in filenames:
#            yield os.path.abspath(os.path.join(dirpath, f))

def absoluteFilePaths(directory):
    for f in os.listdir(directory):
        path = os.path.join(directory, f)
        if os.path.isfile(path):
            yield os.path.abspath(path)

def fair(n,p):
    return max(n-p+1,10)

def bestN(punteggi):
    N = int(0.8*len(results_list))
    return sum(punteggi[:N])

parser = argparse.ArgumentParser(description='Script per calcolare punteggi wooclap')
parser.add_argument('-d','--directory', help='Cartella dove si trovano i risultati di ciascuna gara', default='./')
parser.add_argument('-pp','--pointspolicy', help='Regola di assegnamento punti per ciascuna gara', default='fair', choices=['fair'])
parser.add_argument('-sp','--sumpolicy', help='Regola di somma punti per classifica finale', default='bestN', choices=['bestN'])

args = vars(parser.parse_args())

pp = globals()[args["pointspolicy"]]
sp = globals()[args["sumpolicy"]]

results_list = []
for i in absoluteFilePaths(args["directory"]):
    if not i.startswith("~$") and i.split(".")[-1] in excel_extension:
        print(f'Excel trovato: {i}')
        results_list.append(i)
if results_list == []:
    print("Nessun file excel trovato nella cartella. Specifica la giusta cartella con -d/--directory")
    exit(1)

total_scores = {}
wooclap_scores = {}

for file in results_list:
    results = []
    excel = openpyxl.open(file)
    excel = excel[excel.sheetnames[0]]
    colonna_punteggio = excel.max_column
    for riga in excel.iter_rows(2,excel.max_row-1):
        results.append([riga[4].value,riga[colonna_punteggio-1].value])
        wooclap_scores[riga[4].value] = wooclap_scores.get(riga[4].value,0)+riga[colonna_punteggio-1].value
    results.sort(key=lambda k: k[1], reverse=True)
    for p,r in enumerate(results):
        r[1] = pp(len(results),p+1)
        if r[0] not in total_scores:
            total_scores[r[0]] = []
        total_scores[r[0]].append(r[1])
        total_scores[r[0]].sort(reverse=True)


final_scores = []

for k,v in total_scores.items():
    final_scores.append((k,sp(v)))
final_scores.sort(key=lambda k:(k[1],wooclap_scores[k[0]]),reverse=True)


table = prettytable.PrettyTable()
table.field_names = ["Pos","Nome","Punteggio"]
table.align["Pos"] = "r"
for p,i in enumerate(final_scores):
    table.add_row([p+1, i[0],i[1]])
with open("risultati.txt","w") as f:
    f.write(str(table))
f.close()

print("File risultati.txt creato")
